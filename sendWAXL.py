# import load_workbook
from openpyxl import load_workbook
from docx import Document
import urllib.parse
import webbrowser

def message_text_from_docx_filepath(filepath):
    with open(filepath, 'rb') as file:
        document = Document(file)
    fullText = []
    for para in document.paragraphs:
        fullText.append(para.text)
    return '\n'.join(fullText)

def generate_message_from_file(filepath, name, user, password):
     message = message_text_from_docx_filepath(filepath)
     message_parameters = {"NAME": name, "USER": user, "PASSWORD": password}
     for parameter in message_parameters:
          message = message.replace(parameter, message_parameters[parameter])
     return message

def send_whatsapp(phone_number, message):
     url_encoded_message = urllib.parse.quote(message)
     url = 'https://wa.me/' + phone_number + '?text=' + url_encoded_message
     webbrowser.open_new_tab(url)

def main():
    MESSAGE_FILEPATH = "message.docx"
    SPREADSHEET_FILEPATH = "organized_spreadsheet.xlsx"
    sheet = load_workbook(SPREADSHEET_FILEPATH).active
    MAX_ROW = sheet.max_row

    NAME_COLUMN_INDEX = 1
    USER_COLUMN_INDEX = 2
    PASSWORD_COLUMN_INDEX = 3
    PHONE_COLUMN_INDEX = 4

    for line in range(1, MAX_ROW + 1):
         name = sheet.cell(row=line, column=NAME_COLUMN_INDEX).value
         user = sheet.cell(row=line, column=USER_COLUMN_INDEX).value
         password = sheet.cell(row=line, column=PASSWORD_COLUMN_INDEX).value
         phone_number = str(sheet.cell(row=line, column=PHONE_COLUMN_INDEX).value)

         message = generate_message_from_file(MESSAGE_FILEPATH, name, user, password)
         send_whatsapp(phone_number, message)

         print(phone_number + '\n')

if __name__ == '__main__':
    main()
